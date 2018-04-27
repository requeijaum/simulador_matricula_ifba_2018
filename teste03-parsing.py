#!/usr/bin/env python

# Usar Python3
coding: "utf-8"

# Rascunho do parser de planilhas de horários do Ensino Superior do IFBA
# Rafael F S Requião, Abril/Maio de 2018


# Importar bibliotecas de sempre
import os, sys, io, time, datetime, string
from datetime import date

# Usar interface gráfica com npyscreen
# > http://npyscreen.readthedocs.io/application-structure.html
import npyscreen

# Importar dados da planilha com OpenPyXL
import openpyxl

# Exportar dados... (CSV, JSON, PDF (via TeX)?)
import simplejson as json

#imprimir nome da função - útil para Debug - apenas Python 3
#from __future__ import inspect
import inspect

# persistencia de objetos, para não ficar tendo que importar a planilha sempre
# além de permitir continuar o trabalho incompleto
# https://stackoverflow.com/questions/1773805/how-can-i-parse-a-yaml-file-in-python
import yaml

global main_file
global aux_file

main_file = "db.yaml"	# usar isso pra salvar e carregar turmas
aux_file  = "aux.yaml"  # usar isso pra salvar e carregar cursos, mas futuramente: usar pra horarios e etc.





# GLOBAL AQUI <-----------------------------------
# Definir variaveis
wb 			= None
sheet		= None
sheet_list	= None

turmas = 0
turmas_lista = []

stack_turmas = []

menu = None
curso = ""

#======================================================================
# definir classes

class Turma:
	"Uma classe representando uma turma de alunos com código, nome da disciplina, número da turma, professor, horário (codificado) e número de vagas"
	turmas = 0
	
	#def __call__(self):
	#	return self

	def __init__(self, cod, nome, nturma, nome_prof, horario_cod, vagas): #simplificacao, grafico, link, raizes)

		#lembrar que eu capturo objetos Cell... tenho que acessar os valores!
		
		self.cod 			= cod.value
		self.nome 			= nome.value 		#ajeitar encoding pra unicode ou ansi?
		self.nturma 		= nturma.value
		self.nome_prof 		= nome_prof.value 	#ajeitar encoding pra unicode ou ansi?
		
		
		#e lembrando que horarios e vagas serão lidos e criados em objetos, arrays, sei lá
		#esses não serão objetos Cell, do openpyxl
		
		self.horario_cod 	= horario_cod  # [SEG_INI, TER_INI, QUA_INI, QUI_INI, SEX_INI, SAB_INI]
		self.vagas 			= vagas  # [EVE, MEC, ELE, ADM, POL, RAD, QUI, ADS, MAT, GEO, FIS]
		self.json 			= None
		
		
		Turma.turmas += 1
		#coloquei depois, pq quero que não exista uma Turma 0.
		self.numero			= Turma.turmas
		
	
	def contador(self):
		#print("[DEBUG] Numero de Turmas = " + Turma.turmas)
		return Turma.turmas
	
	
	def horarios_obj(self):
		print("[DEBUG] self.horario_cod = ")
		print(self.horario_cod)
		
		obj = { "seg": self.horario_cod[0], "ter": self.horario_cod[2], "qua": self.horario_cod[4], "qui": self.horario_cod[6], "sex": self.horario_cod[8], "sab": self.horario_cod[10] }
		return obj
		
		
	def vagas_obj(self):
		print("[DEBUG] self.vagas = ")
		print(self.vagas)
		
		obj = {"EVE": self.vagas[0], "MEC": self.vagas[1], "ELE": self.vagas[2], "ADM": self.vagas[3], "POL": self.vagas[4], "RAD": self.vagas[5], "QUI": self.vagas[6], "ADS": self.vagas[7], "MAT": self.vagas[8], "GEO": self.vagas[9], "FIS": self.vagas[10]}
		return obj
		


	def JSON(self):
		#TypeError: Object of type Cell is not JSON serializable
		# parece também que o JSON codifica as coisas em unicode... mas tá tudo OK
		
		#TypeError: Object of type datetime is not JSON serializable
		#https://code-maven.com/serialize-datetime-object-as-json-in-python
		
		#versao velha
		#self.json = json.dumps({"Turma": self.numero, "detalhes": {"cod": self.cod, "nome": self.nome, "nturma": self.nturma, "nome_prof": self.nome_prof, "horario_cod": self.horario_cod, "vagas": self.vagas }}, sort_keys=False, indent=5*" ", default = conv)
		
		#versao nova
		self.json = json.dumps({"Turma": self.numero, "detalhes": {"cod": self.cod, "nome": self.nome, "nturma": self.nturma, "nome_prof": self.nome_prof, "horarios": self.horarios_obj() , "vagas": self.vagas_obj() }}, sort_keys=False, indent=5*" ", default = conv)
		
		
		#print("[DEBUG] type(self.nome) e type(self.nome_prof)  = " + str(type(self.nome)) + "  " + str(type(self.nome_prof)) )
		#print("[DEBUG] JSON = \n" + self.json)
		return self.json

		
#fim de class Questao()

# criar uma classe que vai receber vagas por curso
# outra classe pra receber os códigos de horarios em cada dia da semana

class Menu:
	def __init__(self): #simplificacao, grafico, link, raizes)
		self.horarios 	= ["07:00", "08:40", "10:40", "12:00", "13:20", "15:20", "17:00", "18:40", "20:20", "22:00"]
		
		self.seg = []
		self.ter = []
		self.qua = []
		self.qui = []
		self.sex = []
		self.sab = []
		
		for i in range (0,9):
			self.seg.append(5*" ")
			self.ter.append(5*" ")
			self.qua.append(5*" ")
			self.qui.append(5*" ")
			self.sex.append(5*" ")
			self.sab.append(5*" ")
			
			
			
class PersonalInfo :
	"Um objeto necessário para eu gerir informações pessoais; \
	salvar os cursos existentes na planilha... e talvez manipular como objeto depois!"
	
	#def __call__(self):
	#	return self

	def __init__(self, lista_cursos, meu_curso):
		#vou receber um list com os cursos!
		
		self.lista_cursos		= lista_cursos
		self.meu_curso			= meu_curso
	
	
	#def contador(self):			


# fim da class Cursos()

			
			
# Definir funções auxiliares

from platform import system as system_name # Returns the system/OS name
from os import system as system_call       # Execute a shell command
import glob


def clear_screen():
	os.system('cls' if os.name=='nt' else 'clear')
	
	
	
	
def flush_in():

	#bug introduzido a esta função no quando mudei de:
	#Python 3.4 para Python 3.6.3
	
	#https://docs.python.org/3/library/termios.html	
	'''
	Set the tty attributes for file descriptor fd from the attributes, which is a list like the one returned by tcgetattr(). The when argument determines when the attributes are changed: 
	TCSANOW to change immediately, TCSADRAIN to change after transmitting all queued output, or TCSAFLUSH to change after transmitting all queued output and discarding all queued input.
	'''
	
	#https://linux.die.net/man/3/tcflush
	

	try:
		import msvcrt
		while msvcrt.kbhit():
			msvcrt.getch()
			
	except ImportError:
		import sys, termios, tty 			#inclui modulo "tty"
		#termios.tcflush(sys.stdin, termios.TCIOFLUSH) 	#TCIOFLUSH tá bugando?
		sys.stdin.flush()


def entrar(texto):
	
	#if os.name=='nt':
	#	flush_in() 
	#
	#else:
	#	sys.stdin.flush()
		


	flush_in() 

	return input(texto)
	

# bug nos horários - culpa de quem fez a planilha
# converter "5/4" pra texto... o valor é lido como data!

def conv(o):
	value = None

	if isinstance(o, datetime.datetime):
		value = str(str(o.month) + "/" + str(o.day))

	else:
		value = o
			
  
    	
    	
	return value





def desenharLinha():
	print(80 * "-")

def nomeFuncaoAtual():
	FuncaoAtual = inspect.stack()[1][3] + ", " + inspect.stack()[2][3] + ", " + inspect.stack()[3][3]
	#FuncaoAtual    = inspect.stack()[0][3]
	#[0][3] é a atual! mas n quero essa função... quero a anterior
	#print("[DEBUG] nomeFuncaoAtual() --> " + FuncaoAtual)
	
	return str(FuncaoAtual)
	
	

# funções para persistencia dos dados
#    verificar pq as turmas estão sendo duplicadas - arquivo sendo recopiado pro final

def yamlLoad(filename):
	
	# Read YAML file
	with open(filename, 'r') as stream:
		data_loaded = yaml.load(stream)
		
	return data_loaded


def yamlSave(data, filename):
	# Write YAML file
	with io.open(filename, 'w', encoding='utf8') as outfile:
		yaml.dump(data, outfile, default_flow_style=False, allow_unicode=True)

# fim das funções de persistencia dos dados


#https://stackoverflow.com/questions/5319922/python-check-if-word-is-in-a-string
def contains_word(string, word):
    return f' {word} ' in f' {string} '


def buscarCodigos(query):

	#parece que query não pode ser vazio

	# BUSCAR CODIGOS OBTIDOS
	#nomeFuncaoAtual()
	#print(5*" " + "query = " + str(query) )
	return sortCodigos(query)



def sortCodigos(inicial):

	#chamar função em preLoadAll() ?
	#pra gerar as iniciais possíveis
	
	#eu poderia gerar um objeto Iniciais com atributos de cada inicial?
	
	
	#nomeFuncaoAtual()

	global turmas
	global turmas_lista

	#TypeError: int() argument must be a string, a bytes-like object or a number, not 'Turma'
	#print("[DEBUG] turmas = " , turmas)
	#print(type(turmas))	
	#a = int(turmas)

	#tornar global?	
	iniciais 	= []
	buscarTurma_lista = []
	
	#gerar n importa o que aconteça
	for turma in turmas_lista:
		#print(turmaA)
		codigo = str(turma.cod)[0:3]
		#print(codigo)
		iniciais.append(codigo)
				
	iniciais = remove_duplicates(iniciais)
	#print("[DEBUG]")
	#print(iniciais)	
	#print("\n\n")
	
		
	if inicial == "" or inicial == " " or inicial == "  " or inicial == None :
		#print("[DEBUG] listar todos as iniciais possiveis para os códigos")
		print(5*" " + "Você pode pesquisar pelas seguintes iniciais: \n")
		
		#for i in range(0,len(iniciais)):
		#	print(3*" " + iniciais[i])		#como imprimir na mesma linha?
		print(iniciais)
		#return iniciais


	elif isinstance(inicial, str):

		if len(inicial) == 3:

			for turma in turmas_lista:
				if inicial == str(turma.cod)[0:3]:
					#print(turma.numero)
					buscarTurma_lista.append(turma.numero)
					
			#pegar codigos possiveis cadastrados, a partir do codigo informado
				
			for i in range(0, (turmas-1)): 
				if contains_word(turmas_lista[i].cod, inicial):
					buscarTurma_lista.append(turmas_lista[i].numero)

			#def mostrarPeloIndice(indice): - transformar esse pedaço em uma função auxiliar
	
			for i in range(0, (turmas-1)): 
				for indice in buscarTurma_lista:
					if indice == turmas_lista[i].numero :
						print("\n" + 5*" "+ "Turma: #" + str(turmas_lista[i].numero) + " = " + turmas_lista[i].cod + ", " + turmas_lista[i].nome + " - "+ turmas_lista[i].nturma + "\n" + 20*" " + "(Prof. " + turmas_lista[i].nome_prof + ")")
	
	
			# fim do tradutor de indice pra descricao util
	
			# mudar retorno depois	
			#return t
		
	
	#return (inicial + " ... numeros")



# fim das funcoes auxiliares


# https://www.dotnetperls.com/duplicates-python
def remove_duplicates(values):
    output = []
    seen = set()
    for value in values:
        # If value has not been encountered yet,
        # ... add it to both list and set.
        if value not in seen:
            output.append(value)
            seen.add(value)
    return output



# inicio de funcoes legais

def imprimirInfo():
	#vars globais
	global turmas
	global turmas_lista

	# grab the active worksheet
	ws = wb.active

	#preciso contar quantas turmas estão cadastradas...
	#contar por cada linha
	#pulando a primeira linha
	#e verificando se o codigo são 3 letras e depois 3 numeros
	#fazer verificando por str.len == 6... depois
	
	
	colA = ws["A"]
	turmas = len(colA) -1 #não contar a primeira linha, obviamente
	#print(type(colA))
	print("[DEBUG] numero de turmas cadastradas na planilha (contar por linhas) = " + str(turmas))
	time.sleep(3)
	
	#print(colA)
	#print("\n\n" + "asdf" + "\n\n")
	

	#imprimir cada linha da tabela com suas respectivas informações
	#print("\n ws.rows: \n" + str(tuple(ws.rows)))
	#time.sleep(3)
	#print("\n ws.columns: \n" + str(tuple(ws.columns)))	
	
	#vou criar um tuple e ir atualizando ele (somando)
	#turmas_lista = [] #fazer global?
	
		
	# verificar se a planilha foi editada e mudou a ordem das colunas
	print("[DEBUG] verificar ordem...")
	lista_colunas = []
	for row in ws.iter_rows(min_row=1, max_col=27, max_row=1): #max_row=(turmas+1)
		for cell in row:							#turmas+1 foi pq a row mínima é o numero da linha, e não o número de vezes que iremos descer a linha
			lista_colunas.append(str(cell.value))
	
	print(lista_colunas)
	# devo comparar isso abaixo com o que está na planilha
	
	verify = ['CODIGO', 'None', 'TURMA', 'SEG_INI', 'SEG_FIM', 'TER_INI', 'TER_FIM', 'QUA_INI', 'QUA_FIM', 'QUI_INI', 'QUI_FIM', 'SEX_INI', 'SEX_FIM', 'SAB_INI', 'SAB_FIM', 'EVE', 'MEC', 'ELE', 'ADM', 'POL', 'RAD', 'QUI', 'ADS', 'MAT', 'GEO', 'FIS', 'PROFESSOR']
	
	if (lista_colunas == verify):
		print("[DEBUG] planilha apresenta ordem correta de itens, na linha inicial.")	 
		time.sleep(3)
		
		# fazer a bagaça setar atributos junto com criacao das listas
		# dentro de um loop, com verificador
		#      setattr(Turma[i], 'seg', row[i].value) # Set attribute 'age' at 8
		
		
		
		#colocar essa porra toda dentro do IF	
		for row in ws.iter_rows(min_row=2, max_col=27, max_row=(turmas+1)): #max_row=(turmas+1)
			#for cell in row:							#turmas+1 foi pq a row mínima é o numero da linha, e não o número de vezes que iremos descer a linha
			#	print(str(cell.value))


			#implementar leitura de horarios e vagas
			#verificar como lidar com células vazias - viram "" ?
					
			lista_horario_cod 	= []
			lista_vagas 		= []

		
			#	Turma.horario_cod 	= []  # [SEG_INI, TER_INI, QUA_INI, QUI_INI, SEX_INI, SAB_INI]
			for i in range(3,14):
				if ((type(row[i].value) == type(None) ) ) :
					#lista_horario_cod.append(" ")  				#-> se bugar: restaurar esse
					lista_horario_cod.append("") 					#vazio... pra testar bool(horario_cod[i]) = False
				
				
				else:
					#lista_horario_cod.append(int(row[i].value))	
					#lista_horario_cod.append(row[i].value) 		#-> ainda aparecem datetimes
					#lista_horario_cod.append(conv(row[i].value)) 
					lista_horario_cod.append(converteHorario(conv(row[i].value)))
					 
			
			#	Turma.vagas 			= []  # [EVE, MEC, ELE, ADM, POL, RAD, QUI, ADS, MAT, GEO, FIS]
			# talvez eu precise transformar str() em int(), para os valores --> int(row[x])
			for i in range(15,26):
				if ((type(row[i].value) == type(None) )):
					lista_vagas.append("")
			
				else:
					#lista_vagas.append(int(row[i].value))
					lista_vagas.append(row[i].value)
		
			#ir criando turmas com valores
			#	def __init__(self, cod, nome, nturma, nome_prof, horario_cod, vagas): #simplificacao, grafico, link, raizes)

			#implementar leitura de horario e vagas... a cada busca de linha (row)
			turmas_lista.append(Turma(row[0], row[1], row[2], row[26], lista_horario_cod, lista_vagas)) #time.sleep(1)
		
	
		#fim do for row
	
		print(turmas_lista)
	
		for i in range (0, turmas):
			print(turmas_lista[i].JSON())
	
		#turmas_lista[0].JSON()
		
		#acho q preciso iterar a lista pra capturar todos os objetos... aff
		yamlSave(None) #preciso zerar antes, né		
		yamlSave(turmas_lista)
		
	
	else:
		print("[ERRO] planilha NÃO apresenta ordem correta de itens, na linha inicial.")
		print("Encontrado:  " + lista_colunas)
		print("Deveria ser: " + verify)
	

	#fim do verificador		
	

def decodificaHorario(codigo):

	horario_str = ""

	'''
	0/1 - 13:20 as 15 h
	2/3 - 15:20 as 17 h
	4/5 - 17 as 18:40 h
	6/7 - 18:40 as 20:20 h
	8/9 - 20:20 h as 22 h
	10/11 - 7 as 8:40 h
	12/13 - 8:40 as 10:20 h
	14/15 - 10:40 h as 12:20 h
	'''

	if (codigo == 0 or codigo == "0"):
		horario_str = "13:20"
	
	if (codigo == 1 or codigo == "1"):
		horario_str = "15:00"
	
	if (codigo == 2 or codigo == "2"): #acho que o gap de 20 min. é por causa do intervalo da tarde
		horario_str = "15:20"
	
	if (codigo == 3 or codigo == "3"):
		horario_str = "17:00"
	
	if (codigo == 4 or codigo == "4"):
		horario_str = "17:00" # mesmo que o 3 ?
	
	if (codigo == 5 or codigo == "5"):
		horario_str = "18:40"
			
	if (codigo == 6 or codigo == "6"):
		horario_str = "18:40"
	
	if (codigo == 7 or codigo == "7"):
		horario_str = "20:20"
			
	if (codigo == 8 or codigo == "8"):
		horario_str = "20:20"

	if (codigo == 9 or codigo == "9"):
		horario_str = "22:00" # IFBA SSA fecha neste horários, de seg a sex

	if (codigo == 10 or codigo == "10"):
		horario_str = "07:00" # início das aulas de seg a sab

	if (codigo == 11 or codigo == "11"):
		horario_str = "08:40"

	if (codigo == 12 or codigo == "12"):
		horario_str = "08:40"

	if (codigo == 13 or codigo == "13"):
		horario_str = "10:20"

	if (codigo == 14 or codigo == "14"):
		horario_str = "10:40" # gap de 20 min. pq do intervalo da manhã
	
	if (codigo == 15 or codigo == "15"):
		horario_str = "12:20" # IFBA SSA fecha neste horário aos sábados

	#else:
	#	print(nomeFuncaoAtual())
		
	return horario_str





def converteHorario(codigo):	

	if not isinstance(codigo, int):			#ValueError: invalid literal for int() with base 10: '  '
		codigo = str(codigo) #forçar		#ajeitei conv() pra cuspir ints#ajeitei conv() pra cuspir ints

	
	else :
		codigo = int(codigo)
				

	if isinstance(codigo, int):
		return decodificaHorario(codigo)

	elif isinstance(codigo, str):			#lidar com "5/4" - como interpretar?
		codigo_list = codigo.split("/")
		
		junto = ""
		for numero in codigo_list:
			junto += (decodificaHorario(numero) + " ") 
			
		#return junto[0:11].replace(" ", "/") # "18:40 17:00 " --> "18:40/17:00" 
		return junto[0:11]

	else:
		return codigo


# Definir funções do programa - controle de fluxo de execução

def bemvindo():
	clear_screen()
	flush_in()	
	#choice = ""

	print (30 * "-" , "MENU" , 30 * "-")
	print("\n\n* Simulador de Matrícula - IFBA 2018 *\n\n")

	if "bemvindo" in nomeFuncaoAtual():
		print(5*" " + "*** Dica: aperte \"?\" para ajuda ***\n\n")


	

def selecionarPlanilha():

	print(5*" " + "[DEBUG] selecionarPlanilha()")
	print("\n     Por favor, coloque o arquivo da planilha na mesma pasta onde se encontra o script/executável.\n")

	os.chdir(".")
	index=0
	
	for file in glob.glob("*.xlsx"):
		print("\n" + 5*" "  + str(index) + " " + file)
		index += 1
	
	n_escolhido = entrar("\n\n" + 3*" " + "Escolha o número para a planilha desejada: ")
	
	#print("\n\n" + n_escolhido + "\n\n")
	
	return glob.glob("*.xlsx")[int(n_escolhido)]


def importarPlanilha(valor):

	
	# *** detalhe importante sobre importar planilhas do IFBA ***
	# horarios e vagas podem estar apresentados como data e hora
	# sendo necessário editar a planilha pra apresentar numeros inteiros
	
	# - acho que vou colocar essa mensagem na documentação e na tela de boas vindas
	

	#try:
	if True :

		print("[DEBUG] nome do arquivo da planilha = \n> " + valor)

		global  wb
		#global wb.active
		global  sheet
		global  s
		global  sheet_list

		#chamar planilha

		from openpyxl import Workbook

		wb = openpyxl.load_workbook(valor)

		#metodo novo
		sheet_list = wb.sheetnames # em favor
		print("[DEBUG] sheet_list: " + str(sheet_list))

		#metodo velho
		
		#sheet_list = wb.get_sheet_names() #função já deprecated
		#sheet_list = str.format(str(sheet_list))
		#print("sheet_list: \n" + str(sheet_list))
		
		imprimirInfo()

		
		return True


	#except: 
	#	print("[DEBUG]importarPlanilha() retornou erro!")
	#	return False



	
def mostrarSemana():
	
	global menu
	menu = Menu()
	
	#definir tamanho do console... ou obter tamanho?
	#tamanhoConsole()
	
	#usar tamanho padrão do console = 80x24 pixels
	clear_screen()
	
	#to-do: centralizar
	print(23*" " + "*** UMA SEMANA NORMAL DE AULAS ***") #80-34=46
	print("Dias      Seg       Ter       Qua       Qui       Sex       Sab ")
	print("Horários  ")
	desenharLinha()
	for i in range(0, 9): # 10 horarios possiveis
		print(menu.horarios[i] + 5*" " + "|" + 4*" " + menu.seg[i] + "|" + 4*" " + menu.ter[i] + "|" + 4*" " + menu.qua[i] + "|" + 4*" " + menu.qui[i] + "|" + 4*" " + menu.sex[i] + "|" + 4*" " + menu.sab[i] + "|")
		desenharLinha()
		
		

def listarOpcoes():   #sdds manual de instruções dentro do programa
	clear_screen()
	
	print("L = listaOpcoes()")
	desenharLinha()
	print(5*" " + "I = importarPlanilha()")
	print(5*" " + "M = mostrarSemana()")
	print(5*" " + "B = buscarTurma()")
	print(5*" " + "P = buscarProfessor()")
	print(5*" " + "H = buscarHorario()")
	print(5*" " + "S = mostraStackTurmas")
	desenharLinha()
	
	entrar(3*" " + "Aperte qualquer tecla pra voltar...")



# SEMPRE DENTRO DOS MENUS DE BUSCA - POSSIBILITAR INCLUIR OU EXCLUIR UMA DISCIPLINA

def buscarTurma():
	clear_screen()
	print(nomeFuncaoAtual())
	
	global turmas
	global turmas_lista

	query = entrar(5*" " + "Entre com o código da turma --> ").upper()
	
	buscarCodigos(query)
	
	
	# bug atual
	# Turma: #195 =  EVE 112, EVENTOS TÉCNICOS, CIENTÍFICOS E PROMOCIONAIS - T01
	#	     	     (Prof. GLAURIA JANAINA DOS SANTOS)
	
	buscarTurma_lista = []

	if bool(query): #teste de query vazia!
	
		for i in range(0, (turmas-1)): 
			if contains_word(turmas_lista[i].cod, query):
				buscarTurma_lista.append(turmas_lista[i].numero)
				
		print(buscarTurma_lista)
	
	
		#def mostrarPeloIndice(indice): - transformar esse pedaço em uma função auxiliar
	
		for i in range(0, (turmas-1)): 
			for indice in buscarTurma_lista:
				if indice == turmas_lista[i].numero :
					print("\n" + 5*" "+ "Turma: #" + str(turmas_lista[i].numero) + " = " + turmas_lista[i].cod + ", " + turmas_lista[i].nome + " - "+ turmas_lista[i].nturma + "\n" + 20*" " + "(Prof. " + turmas_lista[i].nome_prof + ")")
		
		# fim do tradutor de indice pra descricao util
	
	
		if buscarTurma_lista != []: 		
			choice = entrar("\n\n" + 5*" " + "Pressione (A)dicionar uma turma; ou qualquer tecla para voltar.\n" + 5*" " + "--> ")
	
			if choice == "a" or choice == "A":
				escolher = entrar(10*" " + "\nDigite o # (número) do índice da turma (exemplo: " + str(buscarTurma_lista[0]) + ")\n" + 15*" " + "--> ")
				adicionarTurma(escolher)

	
	return 0
	
	
	
	
def buscarProfessor():

	clear_screen()
	print(nomeFuncaoAtual())
	
	global turmas
	global turmas_lista
	
	#print("[DEBUG]")
	#print(turmas_lista)
	
	query = entrar(5*" " + "Entre com o nome ou sobrenome do professor --> ")
	query = query.upper()
	
	buscarTurma_lista = []
	
	#a = int(str(turmas))
	#print(a)
	
	for i in range(0, (turmas-1)): 
		#if contains_word(turmas_lista[i].nome_prof, query):
		
		#talvez fazer tratamento pra nomes com acento?
		#por enquanto vou só deixar em maiúsculo forçado
		
		#caso não encontre o nome... por causa de acento
		#AttributeError: 'NoneType' object has no attribute 'upper'
		#adicionei esse if... acho que não tô trabalhando com Data Type corretamente
		
		if turmas_lista[i].nome_prof != None :		
			if contains_word(turmas_lista[i].nome_prof.upper(), query):
				buscarTurma_lista.append(turmas_lista[i].numero)
				
	#print(buscarTurma_lista)
	
	
	#def mostrarPeloIndice(indice): - transformar esse pedaço em uma função auxiliar
	
	for i in range(0, (turmas-1)): 
		for indice in buscarTurma_lista:
			if indice == turmas_lista[i].numero :
				print("\n" + 5*" "+ "Turma: #" + str(turmas_lista[i].numero) + " = " + turmas_lista[i].cod + ", " + turmas_lista[i].nome + " - "+ turmas_lista[i].nturma + "\n" + 20*" " + "(Prof. " + turmas_lista[i].nome_prof + ")")
	
	
	# fim do tradutor de indice pra descricao util
	
	
	if buscarTurma_lista != []: 		
		choice = entrar("\n\n" + 5*" " + "Pressione (A)dicionar uma turma; ou qualquer tecla para voltar.\n" + 5*" " + "--> ")
	
		if choice == "a" or choice == "A":
			escolher = entrar(10*" " + "\nDigite o # (número) do índice da turma (exemplo: " + str(buscarTurma_lista[0]) + ")\n" + 15*" " + "--> ")
			adicionarTurma(escolher)

	
		#for i in range(0, len(buscarTurma_lista)-1): 
		#	if numero in buscarTurma_lista[i]
		#		print("")
	
		# se choice for numero: perguntar se n quer adicionar uma turma? aperte A.
		#elif choice
	
	return 0
	
	
	
	
def buscarHorario():

	global turmas
	global curso
	
	clear_screen()
	print(nomeFuncaoAtual())
	
	buscarTurma_lista = []
	
	#a = int(str(turmas))
	#print(a)
	
	# entrar com dia da semana
	# entrar com horario desejado
	# for loop para verificar - nested for loops
	# mostrar disciplinas, uma vez obtidos os indices
	
	
	#listar TODOS os cursos e horarios possiveis - cadastrados no database yaml
	
	
	#tentativa #1 - busca por horario	
	query = entrar(5*" " + "Entre com o horário desejado (ex: \"18:40\") --> ")
	
	
	for i in range(0, (turmas-1)):
		#preciso que existam: horarios.seg até horarios.sab - criar atributos...
		
		for x in turmas_lista[i].horario_cod:  		#pesquisar cada dia
			
			#print("[DEBUG]")
			#print(turmas_lista[i].horario_cod)

			if isinstance(x, str): # tudo agora é "hh:mm"
					#if contains_word(x, query): 		#query precisar ser string ou int?
					if query == x:
						buscarTurma_lista.append(turmas_lista[i].numero)
									
						
			#ainda dentro de for x in turmas_lista[i].horario_cod:
		
		# ainda dentro de for i in range(0, (turmas-1)):
		
	
	#fim de for i in range(0, (turmas-1)):			
						
	
	#tentativa #2 - busca por dia
	#query = entrar(5*" " + "Entre com o dia desejado --> ")
	#query = entrar(5*" " + "Entre com o horário desejado --> ")
	#for i in range(0, (turmas-1)):
	#	for dia in turmas_lista[i].horario_cod:  		#pesquisar cada dia
	#		for horario in dia:							#pesquisar cada horário em um dia
	#			if contains_word(horario, query):
	#				buscarTurma_lista.append(turmas_lista[i].numero)
	
				
	#print(buscarTurma_lista)
	buscarTurma_lista = remove_duplicates(buscarTurma_lista)
	
	#def mostrarPeloIndice(indice): - transformar esse pedaço em uma função auxiliar
	
	for i in range(0, (turmas-1)): 
		for indice in buscarTurma_lista:
			if indice == turmas_lista[i].numero :
				print("\n" + 5*" "+ "Turma: #" + str(turmas_lista[i].numero) + " = " + turmas_lista[i].cod + ", " + turmas_lista[i].nome + " - "+ turmas_lista[i].nturma + "\n" + 20*" " + "(Prof. " + turmas_lista[i].nome_prof + ")")
	
	
	# fim do tradutor de indice pra descricao util
	
	
	if buscarTurma_lista != []: 		
		choice = entrar("\n\n" + 5*" " + "Pressione (A)dicionar uma turma; ou qualquer tecla para voltar.\n" + 5*" " + "--> ")
	
		if choice == "a" or choice == "A":
			escolher = entrar(10*" " + "\nDigite o # (número) do índice da turma (exemplo: " + str(buscarTurma_lista[0]) + ")\n" + 15*" " + "--> ")
			adicionarTurma(escolher)

	
		#for i in range(0, len(buscarTurma_lista)-1): 
		#	if numero in buscarTurma_lista[i]
		#		print("")
	
		# se choice for numero: perguntar se n quer adicionar uma turma? aperte A.
		#elif choice



	
	return 0



def adicionarTurma(indice):
	#print(nomeFuncaoAtual())
	global stack_turmas
	
	#adicionar turma por indice, quando estiver dentro de uma função de busca
		
	indice = int(indice) - 1  #pq o Turma.numero começa em 1, não em 0, como num List
	
	
	# limitar o stack, de acordo com o número de disciplinas mínimas e máximas 
	# a serem cursada
	# reler as "Normas do Ensino Superior.pdf"
	
	stack_turmas.append(turmas_lista[indice].numero)
	print(10*" " + "Turma adicionada com sucesso!")
	time.sleep(1)
	
	
	
	return 0


def mostrarStackTurmas():
	clear_screen()
	print(nomeFuncaoAtual())

	#listar turmas escolhidas... busca por índice 
	#dentro do meu database (arquivo yaml -> RAM)

	
	global stack_turmas
	print("\n" + "Lista de turmas adicionadas:\n" + 5*" ")
	print(stack_turmas)

	# depois verificar cada turma
	# limitar numero de caracteres do nome da disciplina e do prof?
	
	
	'''
	HUM100 - Filosofia - T01 (Prof. Gracinha)
	MAT200 - Cálculo A. - T02 (Prof. Caribé)
	QUI500 - Química Geral e Tecnológica - T01 (Prof. Miguélito*)
	'''	
	
	
	
	return 0
	

	
	

def menuDebug():
	clear_screen()
	print(nomeFuncaoAtual())
	# debug necessário: 
	
	#transformar cada linha em print() depois
	
	print("01 - arquivo YAML - forçar leitura e escrita")
	print("02 - DUMP - de turmas, seletivo - por numero")
	print("03 - DUMP - de numero de turmas cadastradas")
	print("04 - DUMP - da lista de horários")
	print("05 - DUMP - da lista de códigos equivalente aos horários")
	print("06 - buscarCodigos(query) ")
		
	choice = entrar("\n" + 3*" " + "Selecione sua opção --> ")
	
	
	if choice == "06":
		print(5*" " + "buscarCodigos(query)")
		
		query = entrar(5*" " + "Digite iniciais ou nada --> ")
		print(buscarCodigos(query))
	
	
	
	
def opcoesMenu():
		
		choice = entrar("\n     Selecione a opção desejada (I, M, B, P, H, L, S) --> ")
	
		#verificar se consigo importarPlanilha() e salvar objeto gigante em blob
		#e importar o blob

		if   choice == "i" or choice == "I":
				
			if (importarPlanilha(selecionarPlanilha()) == True) :
				mostrarSemana()
				opcoesMenu()
		
			else:
				print("[DEBUG] importarPlanilha(selecionarPlanilha()) == False")
				time.sleep(3)			
			
		elif choice == "m" or choice == "M":
			mostrarSemana()
			opcoesMenu()
			
			
		elif choice == "b" or choice == "B":
			buscarTurma()
			
			
		elif choice == "p" or choice == "P":
			buscarProfessor()	
		
	
		elif choice == "h" or choice == "H":
			buscarHorario()

		
		elif choice == "l" or choice == "L" or choice == "?":
			listarOpcoes()

		elif choice == "s" or choice == "S":
			mostrarStackTurmas()
		
			
		#elif choice == "a" or choice == "A":
		#	adicionarTurma()

		elif choice == "d" or choice == "D":
			menuDebug()



def preLoadAll():

	global turmas
	global turmas_lista

	try:
		print(5*" " + "Carregando arquivo YAML..." )
		turmas_lista = yamlLoad(main_file)
		print(5*" " + "Arquivo carregado com sucesso!" )
		
		#print("[DEBUG]")
		#print(turmas_lista)
		
		for Turma in turmas_lista:
			turmas += 1
			
		#print(type(turmas))			
		print(5*" " + "Encontram-se cadastradas " + str(turmas) + " turmas." )
		
	
	except FileNotFoundError:
		print(5*" " + "Arquivo " + file + " não encontrado.\n" + 5*" " + "Importe uma planilha para começar.")
	



def perguntarCurso():
	curso = entrar("\n\n" + 10*" " + "Qual o seu curso? (ex: \"ELE\") --> ")
	return curso
	

# Definir função principal
def main():

	#gostaria de implementar handling de KeyboardInterrupt

	if __name__ == '__main__':

		# Escrever minhas funcoes depois:
		bemvindo()
		#flush_in()
		preLoadAll()		
		#flush_in()
		
				
		loop = True
		
		while loop:
			if bool(perguntarCurso()) :	#só liberar funcionalidade após informar o curso
				flush_in()	
				opcoesMenu()
 
  
  
  
# Executar a função principal - parece melhorar velocidade de execução
# > https://twitter.com/jeremybowers/status/984494487371239424

if __name__ == '__main__':
    main()
    